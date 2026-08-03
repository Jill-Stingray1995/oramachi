#!/usr/bin/env python3
"""令和7年国勢調査人口速報集計Excelから人口・面積を全自治体へ適用する。

Usage:
  python3 tools/import-official-statistics-2025.py SOURCE.xlsx [OUTPUT.json]
"""

from __future__ import annotations

import hashlib
import json
import math
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
CITIES_PATH = ROOT / "cities.json"
SOURCE_URL = "https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040454825&fileKind=0"


def plain_name(name: str) -> str:
    return re.sub(r"（[^）]+）$", "", name)


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: import-official-statistics-2025.py SOURCE.xlsx [OUTPUT.json]")
    source_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "research" / "official-statistics-2025.json"
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet = workbook["a01"]
    official: dict[tuple[str, str], dict[str, object]] = {}
    tokyo_aggregate = None
    for row in sheet.iter_rows(min_row=14, values_only=True):
        region_type, pref_value, region_value = row[0], row[1], row[2]
        if pref_value is None or region_value is None:
            continue
        pref_text, region_text = str(pref_value), str(region_value)
        pref = pref_text.split("_", 1)[-1]
        code, name = region_text.split("_", 1) if "_" in region_text else ("", region_text)
        if code == "13100":
            tokyo_aggregate = {"population": int(row[3]), "area_km2": float(row[10])}
            continue
        include = str(region_type) in {"1", "2", "3"} or (
            str(region_type) == "0" and code.startswith("131") and code != "13100"
        )
        if not include:
            continue
        if pref == "高知県" and name == "檮原町":
            name = "梼原町"
        official[(pref, name)] = {
            "code": code,
            "population": int(row[3]),
            "area_km2": float(row[10]),
        }
    if len(official) != 1741:
        raise RuntimeError(f"公式自治体数が1,741ではありません: {len(official)}")

    cities = json.loads(CITIES_PATH.read_text(encoding="utf-8"))
    matched: set[tuple[str, str]] = set()
    changes = []
    for city in cities:
        if city["name"] == "東京":
            if tokyo_aggregate is not None:
                city["stats"] = tokyo_aggregate
            continue
        key = (city["pref"], plain_name(city["name"]))
        record = official.get(key)
        if record is None:
            raise RuntimeError(f"公式統計に自治体がありません: {city['pref']}::{city['name']}")
        matched.add(key)
        before = city.get("stats")
        after = {"population": record["population"], "area_km2": record["area_km2"]}
        if before != after:
            changes.append({"id": f"{city['pref']}::{city['name']}", "before": before, "after": after})
        city["stats"] = after
    if matched != set(official):
        raise RuntimeError(f"未使用の公式自治体があります: {sorted(set(official) - matched)}")

    values = [city for city in cities if city["name"] != "東京"]
    areas = sorted(city["stats"]["area_km2"] for city in values)
    densities = sorted(city["stats"]["population"] / city["stats"]["area_km2"] for city in values)
    low_index = math.floor((len(values) - 1) * 0.2)
    high_index = math.ceil((len(values) - 1) * 0.8)
    thresholds = {
        "areaCompact": areas[low_index],
        "areaLarge": areas[high_index],
        "densityLow": round(densities[low_index], 8),
        "densityHigh": round(densities[high_index], 8),
    }
    records = [
        {
            "id": f"{city['pref']}::{city['name']}",
            "code": official[(city["pref"], plain_name(city["name"]))]["code"],
            **city["stats"],
        }
        for city in values
    ]
    records.sort(key=lambda item: item["code"])
    result = {
        "source": "総務省統計局 令和7年国勢調査 人口速報集計 第1表（人口・令和7年10月1日時点の国土地理院面積）",
        "sourceUrl": SOURCE_URL,
        "referenceDate": "2025-10-01",
        "status": "preliminary",
        "municipalityCount": len(values),
        "populationTotal": sum(city["stats"]["population"] for city in values),
        "areaTotalKm2": round(sum(city["stats"]["area_km2"] for city in values), 2),
        "changedMunicipalities": len(changes),
        "thresholds": thresholds,
        "counts": {
            "popUnder50k": sum(city["stats"]["population"] < 50000 for city in values),
            "pop100k": sum(city["stats"]["population"] >= 100000 for city in values),
            "pop300k": sum(city["stats"]["population"] >= 300000 for city in values),
            "pop500k": sum(city["stats"]["population"] >= 500000 for city in values),
            "areaCompact": sum(city["stats"]["area_km2"] <= thresholds["areaCompact"] for city in values),
            "areaLarge": sum(city["stats"]["area_km2"] >= thresholds["areaLarge"] for city in values),
            "densityLow": sum(city["stats"]["population"] / city["stats"]["area_km2"] <= thresholds["densityLow"] for city in values),
            "densityHigh": sum(city["stats"]["population"] / city["stats"]["area_km2"] >= thresholds["densityHigh"] for city in values),
        },
        "sourceSha256": hashlib.sha256(source_path.read_bytes()).hexdigest(),
        "records": records,
        "changes": changes,
    }
    CITIES_PATH.write_text(json.dumps(cities, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({key: result[key] for key in ("municipalityCount", "populationTotal", "changedMunicipalities", "thresholds", "counts")}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

